"""
ml_trainer.py — Training pipeline for the section routing classifier
====================================================================
Orchestrates:
  1. Parse actual plan files to extract ground-truth labels
  2. Match labels to WIP coil attributes
  3. Generate synthetic labels from rule engine for unlabelled coils
  4. Train / retrain the XGBoost classifier
  5. Evaluate and report performance

CLI:
    python ml_trainer.py train   --wip WIP.xlsx --actual ACTUAL.xlsx --model model.pkl
    python ml_trainer.py retrain --model model.pkl --wip WIP.xlsx --actual ACTUAL.xlsx
    python ml_trainer.py eval    --model model.pkl --wip WIP.xlsx --actual ACTUAL.xlsx
    python ml_trainer.py report  --model model.pkl
"""

import argparse
import os
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(__file__))

from generator     import load_wip, filter_rolling_coils
from ml_classifier import SectionClassifier, LABEL_TO_IDX
from parser        import normalise_header


# ── Helpers ──────────────────────────────────────────────────────────────────

def _parse_section_header(header: str) -> str:
    """Convert a raw section header to 'SECTION_KEY|MILL' label."""
    sec_key, mill = normalise_header(header)
    if sec_key == 'UNKNOWN':
        return None
    # Normalise mill
    mill_map = {
        'CRM04':    'CRM04',
        'CRM06':    'CRM06',
        'CRM04/06': 'CRM04',   # default combined to CRM04 for label
    }
    mill = mill_map.get(mill, mill)
    return f"{sec_key}|{mill}"


def extract_labels_from_actual(actual_path: str,
                                sheet_name: str = None) -> dict:
    """
    Parse an actual plan Excel file and return
    {coil_number: 'SECTION_KEY|MILL'} dict.
    """
    wb = openpyxl.load_workbook(actual_path, data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    labels = {}
    current_header = None

    for row in ws.iter_rows(values_only=True):
        if not any(v for v in row if v is not None):
            continue
        first = str(row[0]).strip() if row[0] else ''
        non_empty = [v for v in row if v is not None]

        # Section header detection
        if (any(kw in first.upper() for kw in
                ['ROLLING','FINISH','SKIN','TUBE FH'])
                and len(non_empty) <= 3):
            current_header = first
            continue

        if current_header is None:
            continue

        # Data row
        try:
            wt   = float(row[5]) if row[5] else 0
            coil = str(row[1]).strip() if row[1] else ''
            if wt >= 0.5 and coil and coil not in ('Batch', 'nan', 'None'):
                label = _parse_section_header(current_header)
                if label and label in LABEL_TO_IDX:
                    labels[coil] = label
        except (TypeError, ValueError):
            continue

    return labels


# ── Training pipeline ────────────────────────────────────────────────────────

def train_from_pair(wip_path: str, actual_path: str,
                    model_path: str,
                    existing_model: SectionClassifier = None,
                    sheet_name: str = None,
                    verbose: bool = True) -> SectionClassifier:
    """
    Train (or incrementally retrain) from one WIP + actual plan pair.

    Steps:
      1. Load WIP, filter eligible coils
      2. Extract ground-truth labels from actual plan
      3. Add actual-labelled coils (3x weight)
      4. Add rule-engine synthetic labels for remaining coils (1x weight)
      5. Train / retrain
      6. Save model
    """
    if verbose:
        print(f"\n── Loading WIP: {Path(wip_path).name}")
    raw  = load_wip(wip_path)
    eli  = filter_rolling_coils(raw)

    if verbose:
        print(f"   Eligible coils: {len(eli)}")

    # Extract ground-truth labels
    if verbose:
        print(f"── Parsing actual plan: {Path(actual_path).name}")
    labels = extract_labels_from_actual(actual_path, sheet_name)
    if verbose:
        print(f"   Ground-truth labels: {len(labels)}")
        from collections import Counter
        label_counts = Counter(labels.values())
        for lbl, n in label_counts.most_common():
            print(f"     {n:3d}  {lbl}")

    # Load or create classifier
    clf = existing_model or SectionClassifier()
    if existing_model is None and Path(model_path).exists():
        clf.load(model_path)
        if verbose:
            print(f"── Loaded existing model ({clf.n_samples} samples)")

    # Add actual labels (high weight)
    n_actual = clf.add_training_data(eli, labels, source='actual')
    if verbose:
        print(f"── Added {n_actual} actual-labelled samples")

    # Add synthetic labels for coils not in actual plan
    labelled_coils = set(labels.keys())
    unlabelled = eli[~eli['Coil Number'].isin(labelled_coils)]
    n_synthetic = clf.add_synthetic_data(unlabelled)
    if verbose:
        print(f"── Added {n_synthetic} synthetic-labelled samples")

    # Train
    result = clf.train(verbose=verbose)

    # Save
    clf.save(model_path)
    if verbose:
        print(f"── Model saved: {model_path}")

    return clf


def evaluate(clf: SectionClassifier,
             wip_path: str,
             actual_path: str,
             sheet_name: str = None) -> dict:
    """
    Evaluate classifier accuracy against a held-out actual plan.
    Returns accuracy metrics dict.
    """
    raw  = load_wip(wip_path)
    eli  = filter_rolling_coils(raw)
    labels = extract_labels_from_actual(actual_path, sheet_name)

    # Only evaluate on coils that appear in actual plan
    eval_df = eli[eli['Coil Number'].isin(labels)]
    if eval_df.empty:
        return {'error': 'No matching coils found'}

    preds = clf.predict_batch(eval_df)
    pred_dict = {row['coil_number']: f"{row['section']}|{row['mill']}"
                 for _, row in preds.iterrows()}

    correct = 0
    wrong   = []
    low_conf= []

    for coil, true_label in labels.items():
        if coil not in pred_dict:
            continue
        pred_label = pred_dict[coil]
        conf = preds[preds['coil_number']==coil]['confidence'].values[0]
        if pred_label == true_label:
            correct += 1
        else:
            wrong.append({
                'coil':       coil,
                'true':       true_label,
                'predicted':  pred_label,
                'confidence': round(conf, 3),
            })
        if conf < 0.70:
            low_conf.append({'coil': coil, 'label': true_label,
                             'confidence': round(conf, 3)})

    n_eval = len([c for c in labels if c in pred_dict])
    acc    = correct / n_eval if n_eval else 0

    return {
        'n_evaluated':    n_eval,
        'n_correct':      correct,
        'accuracy':       round(acc, 4),
        'n_wrong':        len(wrong),
        'n_low_conf':     len(low_conf),
        'wrong_details':  wrong,
        'low_conf_coils': low_conf,
    }


def print_report(clf: SectionClassifier) -> None:
    """Print a human-readable summary of the model state."""
    print("=" * 60)
    print("  ML CLASSIFIER — MODEL REPORT")
    print("=" * 60)
    print(f"  Trained        : {clf.is_trained}")
    print(f"  Total samples  : {clf.n_samples}")
    actual = sum(1 for s in clf._sources if s == 'actual')
    synth  = sum(1 for s in clf._sources if s == 'synthetic')
    print(f"  Actual labels  : {actual}")
    print(f"  Synthetic labels: {synth}")
    print(f"  Ready to use   : {clf.ready}")

    if clf.training_log:
        last = clf.training_log[-1]
        print(f"\n  Last training  : {last.get('trained_at','?')[:10]}")
        if last.get('cv_accuracy'):
            print(f"  CV Accuracy    : {last['cv_accuracy']*100:.1f}%")

    if clf.is_trained:
        print("\n  TOP FEATURES:")
        fi = clf.feature_importance(top_n=10)
        for _, row in fi.iterrows():
            bar = '█' * int(row['importance'] * 100)
            print(f"    {row['feature']:25s} {bar} {row['importance']:.4f}")
    print("=" * 60)


# ── CLI ───────────────────────────────────────────────────────────────────────

def cmd_train(args):
    clf = train_from_pair(
        wip_path    = args.wip,
        actual_path = args.actual,
        model_path  = args.model,
        verbose     = True,
    )
    print_report(clf)


def cmd_retrain(args):
    """Incrementally add new data to existing model."""
    clf = SectionClassifier()
    if not clf.load(args.model):
        print(f"No existing model at {args.model} — creating fresh.")
    clf = train_from_pair(
        wip_path       = args.wip,
        actual_path    = args.actual,
        model_path     = args.model,
        existing_model = clf,
        verbose        = True,
    )
    print_report(clf)


def cmd_eval(args):
    clf = SectionClassifier()
    if not clf.load(args.model):
        print(f"ERROR: No model found at {args.model}")
        sys.exit(1)
    result = evaluate(clf, args.wip, args.actual)
    print(f"\n{'═'*55}")
    print(f"  EVALUATION RESULTS")
    print(f"{'═'*55}")
    print(f"  Coils evaluated : {result['n_evaluated']}")
    print(f"  Correct         : {result['n_correct']}")
    print(f"  Accuracy        : {result['accuracy']*100:.1f}%")
    print(f"  Wrong           : {result['n_wrong']}")
    print(f"  Low confidence  : {result['n_low_conf']} (<70%)")
    if result['wrong_details']:
        print(f"\n  WRONG PREDICTIONS:")
        for w in result['wrong_details']:
            print(f"    {w['coil']} | true={w['true']:30s} "
                  f"pred={w['predicted']:30s} conf={w['confidence']:.2f}")
    print(f"{'═'*55}")


def cmd_report(args):
    clf = SectionClassifier()
    if not clf.load(args.model):
        print(f"ERROR: No model found at {args.model}")
        sys.exit(1)
    print_report(clf)


def build_parser():
    p = argparse.ArgumentParser(
        prog='ml_trainer.py',
        description='ML section routing classifier — training pipeline')
    sub = p.add_subparsers(dest='command', required=True)

    t = sub.add_parser('train', help='Train from WIP + actual plan pair')
    t.add_argument('--wip',    required=True)
    t.add_argument('--actual', required=True)
    t.add_argument('--model',  default='models/section_clf.pkl')

    r = sub.add_parser('retrain',
                       help='Incrementally add new data to existing model')
    r.add_argument('--wip',    required=True)
    r.add_argument('--actual', required=True)
    r.add_argument('--model',  default='models/section_clf.pkl')

    e = sub.add_parser('eval', help='Evaluate model on a held-out plan')
    e.add_argument('--wip',    required=True)
    e.add_argument('--actual', required=True)
    e.add_argument('--model',  default='models/section_clf.pkl')

    rp = sub.add_parser('report', help='Print model summary')
    rp.add_argument('--model', default='models/section_clf.pkl')

    return p


if __name__ == '__main__':
    parser = build_parser()
    args   = parser.parse_args()
    dispatch = {
        'train':   cmd_train,
        'retrain': cmd_retrain,
        'eval':    cmd_eval,
        'report':  cmd_report,
    }
    dispatch[args.command](args)
