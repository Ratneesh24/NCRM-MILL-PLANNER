"""
parser.py — Read a formatted mill plan .xlsx back into a structured dict.

Both the generated plan and the planner-corrected actual plan are in the same
format, so one parser handles both.

Output structure:
    {
      "date": <date>,
      "sections": [
          {
            "raw_header": "<full header text>",
            "section_key": "<normalised key, e.g. HT_FINISH>",
            "mill":        "<CRM04 | CRM06 | CRM04/06>",
            "coils": [
                { col_name: value, ... },   # one dict per data row
                ...
            ]
          },
          ...
      ]
    }
"""

import re
from datetime import date, datetime

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Section header normalisation
# ---------------------------------------------------------------------------
# Maps keyword patterns in the header to (section_key, mill_code)
_HEADER_PATTERNS = [
    (r'ROLLING ON BRIGHT ROLLS.*CRM.?04',          'ROLLING_BRIGHT',         'CRM04'),
    (r'FIRST ROLLING ON LIGHT MATT ROLLS.*CRM.?06','FIRST_ROLLING',          'CRM06'),
    (r'RE.?ROLLING ON LIGHT MATT ROLLS.*CRM.?06',  'RE_ROLLING',             'CRM06'),
    (r'R/R ROLLING ON LIGHT MATT ROLLS.*CRM.?06',  'RE_ROLLING',             'CRM06'),
    (r'H&T FINISH ON BRIGHT ROLLS.*CRM.?04',       'HT_FINISH',              'CRM04'),
    (r'H&T FINISH ON BRIGHT ROLLS.*CRM.?06',       'HT_FINISH',              'CRM06'),
    # CRCA finish — check CRM06 before CRM04 (LG Bala)
    (r'CRCA FIN.*ON BRIGHT ROLLS.*CRM.?06',        'CRCA_FINISH_CRM06',      'CRM06'),
    (r'CRCA FIN.*ON BRIGHT ROLLS.*CRM.?04',        'CRCA_FINISH',            'CRM04'),
    # Skin pass variants (match loosely to absorb planner typos)
    (r'SKIN.?PASS ON SUPER.?BRIGHT ROLLS.*CRM.?04','SKIN_PASS_SUPER_BRIGHT', 'CRM04'),
    (r'SKIN.?PASS ON CHROME.?PLAT.*CRM.?04',       'SKIN_PASS_CHROME',       'CRM04'),
    (r'TUBE FH.*ON BRIGHT ROLLS.*CRM.?04.?06',     'TUBE_FH',                'CRM04/06'),
    (r'TUBE FH.*ON BRIGHT ROLLS.*CRM.?06',         'TUBE_FH',                'CRM06'),
    (r'TUBE FH.*ON BRIGHT ROLLS.*CRM.?04',         'TUBE_FH',                'CRM04'),
    (r'SKIN.?PASS ON H.*MATT ROLLS.*CRM.?06',      'SKIN_PASS_HEAVY_MATT',   'CRM06'),
    # Rolling on light matt — most permissive, keep last
    (r'ROLLING ON LIGHT MATT ROLLS.*CRM.?04.?06',  'ROLLING',                'CRM04/06'),
    (r'ROLLING ON LIGHT MATT ROLLS.*CRM.?04',      'ROLLING',                'CRM04'),
    (r'ROLLING ON LIGHT MATT ROLLS.*CRM.?06',      'ROLLING',                'CRM06'),
    # Fallback: anything with CRM mentions
    (r'ROLLING.*CRM.?04.?06',                      'ROLLING',                'CRM04/06'),
    (r'ROLLING.*CRM.?04',                          'ROLLING',                'CRM04'),
    (r'ROLLING.*CRM.?06',                          'ROLLING',                'CRM06'),
]


def normalise_header(raw):
    """
    Given a raw section header string, return (section_key, mill_code).
    Returns ('UNKNOWN','UNKNOWN') if no pattern matches.
    """
    if not raw:
        return 'UNKNOWN', 'UNKNOWN'
    text = raw.upper()
    for pattern, sec_key, mill in _HEADER_PATTERNS:
        if re.search(pattern, text):
            return sec_key, mill
    return 'UNKNOWN', 'UNKNOWN'


# ---------------------------------------------------------------------------
# Column name mapping for reading back a plan sheet
# ---------------------------------------------------------------------------
_COL_MAP = {
    'date':                  'Date',
    'batch':                 'Batch',
    'planning/ so no.':      'SO No',
    'planning/so no.':       'SO No',
    'so no':                 'SO No',
    'thick':                 'Actual Thick',
    'width':                 'Actual Width',
    'weight':                'Input Coil Weight',
    'rt':                    'Plan Rolling Thick 1',
    'customer':              'Customer Desc',
    'prod.code':             'Product Code',
    'quality code':          'Actual Quality',
    'tdc no':                'Cust TDC',
    'plant':                 'Production Plant',
    'staorage loc':          'Storage Location',
    'storage loc':           'Storage Location',
    'planning reamrk':       'Planning Remark',
    'planning remark':       'Planning Remark',
    'current work center':   'Current Stage',
    'next work center':      'Next Stage',
    'route':                 'Process Route',
    'finish':                'Surface Finish',
    'age':                   'Coil Age(# Days)',
}


def _map_col(raw_header):
    return _COL_MAP.get(str(raw_header).strip().lower(), str(raw_header).strip())


# ---------------------------------------------------------------------------
# Row classification helpers
# ---------------------------------------------------------------------------
def _is_section_header(row_values):
    """
    A row is a section header if it has very few non-empty cells
    and the first non-empty cell contains keywords.
    """
    non_empty = [v for v in row_values if v is not None and str(v).strip() != '']
    if len(non_empty) == 0:
        return False
    first = str(non_empty[0]).upper()
    keywords = ['ROLLING', 'FINISH', 'SKIN', 'TUBE FH', 'PLANNING FOR']
    return any(kw in first for kw in keywords) and len(non_empty) <= 3


def _is_column_header(row_values):
    """The column header row starts with 'Date' or 'Batch' in first two cells."""
    joined = ' '.join(str(v) for v in row_values[:4] if v is not None).upper()
    return 'DATE' in joined and ('BATCH' in joined or 'THICK' in joined)


def _is_subtotal_or_blank(row_values, col_count):
    """Subtotal rows have very few values (< 3 non-empty) after position 3."""
    non_empty = [v for v in row_values if v is not None and str(v).strip() not in ('', '0')]
    return len(non_empty) <= 2


def _is_data_row(row_values, col_indices):
    """A data row has a numeric value in the Weight column position."""
    weight_idx = col_indices.get('Input Coil Weight', -1)
    if weight_idx < 0 or weight_idx >= len(row_values):
        return False
    w = row_values[weight_idx]
    try:
        return float(w) >= 0.1
    except (TypeError, ValueError):
        return False


# ---------------------------------------------------------------------------
# Main parse function
# ---------------------------------------------------------------------------
def parse_plan(filepath, target_date=None):
    """
    Parse a mill plan workbook.

    Parameters
    ----------
    filepath    : path to the .xlsx file
    target_date : datetime.date (optional).  If supplied, only that sheet is
                  parsed.  If None, the first sheet is parsed.

    Returns
    -------
    dict with keys: "date", "sections"
    """
    wb = load_workbook(filepath, data_only=True)

    # Select the right sheet
    if target_date is not None:
        sheet_name = target_date.strftime('%d-%m-%Y')
        if sheet_name not in wb.sheetnames:
            # Try to find by partial match
            matches = [s for s in wb.sheetnames if target_date.strftime('%d') in s]
            sheet_name = matches[0] if matches else wb.sheetnames[0]
    else:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # Try to read date from sheet name
    parsed_date = None
    try:
        parsed_date = datetime.strptime(sheet_name, '%d-%m-%Y').date()
    except ValueError:
        parsed_date = target_date or date.today()

    rows = list(ws.iter_rows(values_only=True))

    # Find column header row and build index
    col_indices = {}
    header_row_idx = None
    for i, row in enumerate(rows):
        if _is_column_header(row):
            header_row_idx = i
            for j, cell_val in enumerate(row):
                mapped = _map_col(cell_val)
                col_indices[mapped] = j
            break

    if header_row_idx is None:
        # Fallback: use positional mapping (matches OUTPUT_COLUMNS order)
        _fallback_cols = [
            'Date', 'Batch', 'SO No', 'Actual Thick', 'Actual Width',
            'Input Coil Weight', 'Plan Rolling Thick 1', 'Customer Desc',
            'Product Code', 'Actual Quality', 'Cust TDC', 'Production Plant',
            'Storage Location', 'Planning Remark', 'Current Stage',
            'Next Stage', 'Process Route', 'Surface Finish', 'Coil Age(# Days)',
        ]
        col_indices = {c: i for i, c in enumerate(_fallback_cols)}
        header_row_idx = 1   # assume row 2 is header (0-based = 1)

    # Walk rows after the header
    sections = []
    current_section = None

    for row in rows[header_row_idx + 1:]:
        # Skip the merged cells that have None propagated across
        non_empty = [v for v in row if v is not None and str(v).strip() != '']
        if not non_empty:
            continue

        first_val = str(row[0]).strip() if row[0] is not None else ''

        # Section header detection: single long merged cell with keywords
        if _is_section_header(row):
            sec_key, mill = normalise_header(first_val)
            current_section = {
                'raw_header': first_val,
                'section_key': sec_key,
                'mill': mill,
                'coils': [],
            }
            sections.append(current_section)
            continue

        # Skip subtotal / grand total / planning block rows
        if _is_subtotal_or_blank(row, len(col_indices)):
            continue

        # Data row
        if current_section is None:
            continue

        if not _is_data_row(row, col_indices):
            continue

        coil_dict = {}
        for col_name, idx in col_indices.items():
            if idx < len(row):
                coil_dict[col_name] = row[idx]
            else:
                coil_dict[col_name] = None

        # Add the raw coil number from SO No or construct a key
        so_no = str(coil_dict.get('SO No', '')).strip()
        coil_dict['coil_number'] = so_no    # best available identifier

        current_section['coils'].append(coil_dict)

    return {
        'date':     parsed_date,
        'sections': sections,
    }


def build_coil_index(parsed_plan):
    """
    Return {coil_identifier: {section, position, data}} for diff engine.
    Identifier = SO No (best proxy for coil identity in the plan).
    """
    index = {}
    for sec in parsed_plan['sections']:
        for pos, coil in enumerate(sec['coils']):
            key = coil.get('coil_number') or coil.get('SO No', '')
            key = str(key).strip()
            if not key:
                continue
            index[key] = {
                'section': sec['section_key'],
                'mill':    sec['mill'],
                'position': pos,
                'header':   sec['raw_header'],
                'data':     coil,
            }
    return index
