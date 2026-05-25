"""
Plan generator.

Orchestrates: load WIP → filter eligible coils → assign sections →
sort within sections → write formatted Excel.
"""

import os
from datetime import date, datetime, timedelta

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from constants import (
    OUTPUT_COLUMNS, N_COLS, SECTION_ORDER, MILL_ORDER,
    SECTION_COLOURS, SECTION_SHORT_NAME, CRM04_PRIORITY, CRM06_PRIORITY,
    build_section_label, abbreviate_customer,
)
from sectioning import assign_section_with_learning, _f, _s


EXCEL_EPOCH = date(1899, 12, 30)
THIN = Side(border_style='thin', color='808080')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Load + filter
# ---------------------------------------------------------------------------
REQUIRED_COLUMNS = [
    'Actual Thick', 'Actual Width', 'Input Coil Weight',
    'Plan Rolling Thick 1', 'Current Stage', 'Next Stage',
    'Product Code', 'Actual Quality', 'Cust TDC', 'Production Plant',
]

def _detect_sheet(filepath):
    """
    Return the best sheet name to use from the workbook.
    Preference order:
      1. 'Sheet1'  (exact match)
      2. Any sheet whose name contains 'sheet' (case-insensitive)
      3. First sheet in the workbook
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    if 'Sheet1' in names:
        return 'Sheet1'
    for n in names:
        if 'sheet' in n.lower():
            return n
    return names[0]   # fall back to first sheet


def load_wip(filepath):
    """
    Read the WIP file and normalise dtypes.

    - Accepts any filename.
    - Auto-detects the correct sheet (prefers 'Sheet1', falls back gracefully).
    - Raises a clear ValueError listing missing columns if the structure
      does not match expectations, instead of a cryptic KeyError later.
    """
    sheet_name = _detect_sheet(filepath)
    df = pd.read_excel(filepath, sheet_name=sheet_name)

    # Strip whitespace from column headers (common when exported from SAP)
    df.columns = df.columns.str.strip()

    # Validate required columns
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"The uploaded file is missing {len(missing)} required column(s):\n"
            + "\n".join(f"  • {c}" for c in missing)
            + "\n\nMake sure you are uploading the correct WIP coil staging export "
              f"(sheet used: '{sheet_name}', columns found: {list(df.columns[:8])} …)"
        )

    # Coerce numerics that we rely on
    num_cols = ['Actual Thick', 'Actual Width', 'Input Coil Weight',
                'Plan Rolling Thick 1', 'Coil Age(# Days)', 'Production Plant',
                'Cust Thick', 'Plan Weight', 'Balance Coil Weight']
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce')

    # Strip whitespace on string cols
    str_cols = ['Coil Number', 'SO No', 'Customer Desc', 'Product Code',
                'Actual Quality', 'Cust TDC', 'Storage Location',
                'Planning Remark', 'Current Stage', 'Next Stage',
                'Process Route', 'Edge', 'Surface Finish', 'Work Center',
                'Last Production Stage']
    for c in str_cols:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()
            df[c] = df[c].replace({'nan': '', 'NaN': '', 'None': ''})

    return df


# Excluded next-stages — coil is past rolling, going to FG/dispatch
EXCLUDED_NEXT = {
    '11-FG', 'FG/Palletization', 'PACK', 'PALLETIZATION',
    'IS-INSPECTION TABLE/CTL', 'N-GRINDING', 'E-EDGE ROUNDING',
    'CT-COLOR TEMPERING', 'T-HR SLITTER',
}


def filter_rolling_coils(df, learning_db=None):
    """
    Apply inclusion criteria.
    STRICT RULE: only coils whose Current Stage is exactly 'ROLLING MILL'.
    No other stage is included regardless of Last Production Stage.
    """
    # Production plant filter
    if 'Production Plant' in df.columns:
        df = df[df['Production Plant'].fillna(0).astype(int).isin([760, 761])]

    # STRICT: Current Stage must be exactly ROLLING MILL
    df = df[df['Current Stage'] == 'ROLLING MILL'].copy()

    # Drop excluded next stages
    df = df[~df['Next Stage'].isin(EXCLUDED_NEXT)]

    # HC80 coils: exclude if Actual Thick < Plan Rolling Thick 1
    # (target not yet reached — planner defers these to next campaign)
    # Include if Actual Thick == Plan Rolling Thick 1 (at target, ready for anneal)
    hc80_below_target = ((df['Cust TDC'] == 'HC80') &
                         (df['Actual Thick'].fillna(0) <
                          df['Plan Rolling Thick 1'].fillna(0) - 0.05))
    df = df[~hc80_below_target]

    # HC80 in transit storage (RNM6 etc.) also exclude
    hc80_rnm = ((df['Cust TDC'] == 'HC80') &
                (df['Storage Location'].isin(['RNM6','RNM4','RNM5'])))
    df = df[~hc80_rnm]

    # TATFHC/RC01 = new arrivals not yet in campaign
    # Exception: if PP-PENDING, it IS in the campaign (SAP lag) — include it
    tatfhc_rc01_new = ((df['Actual Quality'] == 'TATFHC') &
                       (df['Storage Location'] == 'RC01') &
                       (~df['Next Stage'].str.upper().str.contains('PP-PENDING', na=False)))
    df = df[~tatfhc_rc01_new]

    # PP-PENDING FOR PLAN: exclude UNLESS it is a TATFHC/TR17 Tube FH coil
    # — planners always include Tube FH coils regardless of PP-PENDING status
    # because they are physically ready and roll campaign must be continuous.
    pp_mask    = df['Next Stage'].str.upper().str.contains('PP-PENDING', na=False)
    tube_fh    = (df['Actual Quality'] == 'TATFHC') & (df['Product Code'] == 'C09')
    df = df[~pp_mask | tube_fh]

    # Drop very-low-weight stubs
    df = df[df['Input Coil Weight'].fillna(0) >= 0.5]

    # Drop coils with no rolling target
    # Exception: TATFHC/TR17 Tube FH with RT=0 but Next=R-C R SLITTER
    # — planner assigns RT manually on the floor, include these
    tube_fh_no_rt = ((df['Actual Quality'] == 'TATFHC') &
                     (df['Product Code'] == 'C09') &
                     (df['Next Stage'] == 'R-C R SLITTER'))
    df = df[(df['Plan Rolling Thick 1'].fillna(0) > 0) | tube_fh_no_rt]

    # Drop HOLD coils
    remark = df['Planning Remark'].fillna('').astype(str).str.lower()
    # Only exclude if 'hold' appears as a standalone word (not embedded in route string)
    # e.g. 'on hold' or 'hold' alone — NOT '>>hold' which is a route thickness step
    import re as _re
    hold_standalone = remark.apply(
        lambda r: bool(_re.search(r'(?<![>0-9])hold(?![>0-9])', r)))
    df = df[~hold_standalone]
    df = df[df['Next Stage'].astype(str).str.upper() != 'HOLD']

    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Assign + group + sort
# ---------------------------------------------------------------------------
def assign_all(df, learning_db=None):
    """Return df with new columns: _section, _mill, _source."""
    sections, mills, sources = [], [], []
    for _, row in df.iterrows():
        s, m, src = assign_section_with_learning(row, learning_db)
        sections.append(s)
        mills.append(m)
        sources.append(src)
    df = df.copy()
    df['_section'] = sections
    df['_mill']    = mills
    df['_source']  = sources
    return df


def split_combined_mill(group_df, section_key):
    """
    Split any remaining CRM04/06 combined groups into CRM04 and CRM06.
    Uses Work Center first, then width-based heuristic.
    Never returns CRM04/06 combined — always splits into two.
    """
    mill = group_df['_mill'].iloc[0] if not group_df.empty else 'CRM04'

    # If already a specific mill, just return it
    if mill != 'CRM04/06':
        return {mill: group_df}

    # Try Work Center split first
    wc = group_df['Work Center'].fillna('').astype(str).str.upper()
    crm04_wc = wc.str.contains('CRM04|SNCRM04|SWPL00')
    crm06_wc = wc.str.contains('CRM06|SWCRS1|SWCRS2')

    has04 = crm04_wc.any()
    has06 = crm06_wc.any()

    if has04 and has06:
        rest = group_df[~crm04_wc & ~crm06_wc]
        g04  = pd.concat([group_df[crm04_wc], rest]).drop_duplicates()
        g06  = group_df[crm06_wc]
        return {'CRM04': g04.copy(), 'CRM06': g06.copy()}

    if has04:
        return {'CRM04': group_df.copy()}
    if has06:
        return {'CRM06': group_df.copy()}

    # Width-based heuristic split:
    # Wider / thicker coils → CRM04 (higher-capacity mill)
    # Narrower / thinner   → CRM06
    if section_key == 'TUBE_FH':
        # RP01/RP02 storage → CRM04; R037 narrow → CRM06
        stor = group_df['Storage Location'].fillna('').astype(str)
        crm04_stor = stor.isin(['RP01', 'RP02', 'R032'])
        g04 = group_df[crm04_stor]
        g06 = group_df[~crm04_stor]
        if not g04.empty and not g06.empty:
            return {'CRM04': g04.copy(), 'CRM06': g06.copy()}
        if not g04.empty:
            return {'CRM04': g04.copy()}
        return {'CRM06': g06.copy()}

    # General: split by thickness — thicker heavy gauge → CRM04
    thick = group_df['Actual Thick'].fillna(0).astype(float)
    median_thick = thick.median()
    g04 = group_df[thick >= median_thick]
    g06 = group_df[thick < median_thick]
    if g04.empty:
        return {'CRM06': group_df.copy()}
    if g06.empty:
        return {'CRM04': group_df.copy()}
    return {'CRM04': g04.copy(), 'CRM06': g06.copy()}


def sort_section(section_df):
    """Apply the width-cascade sort: width↓, thick↓, SO↑, age↓, weight↓."""
    df = section_df.copy()
    df['_so'] = df['SO No'].fillna('').astype(str)
    df = df.sort_values(
        by=['Actual Width', 'Actual Thick', '_so',
            'Coil Age(# Days)', 'Input Coil Weight'],
        ascending=[False, False, True, False, False],
        kind='mergesort',
    )
    return df.drop(columns=['_so'])


def build_sections(df, learning_db=None):
    """
    Returns ordered list: [{section_key, mill, label, coils_df}].
    Skips empty sections, splits combined-mill sections where appropriate.
    """
    df = df[df['_section'] != 'OTHER']

    out = []
    for section_key in SECTION_ORDER:
        sec_df = df[df['_section'] == section_key]
        if sec_df.empty:
            continue

        # First resolve any remaining CRM04/06 combined assignments
        resolved = []
        for _, row in sec_df.iterrows():
            if row['_mill'] == 'CRM04/06':
                # Re-run split logic on this single row's group
                sub = split_combined_mill(
                    sec_df[sec_df['_mill'] == 'CRM04/06'], section_key)
                for sub_mill, sub_df in sub.items():
                    sub_df = sub_df.copy()
                    sub_df['_mill'] = sub_mill
                    resolved.append(sub_df)
                break
            else:
                resolved.append(sec_df[sec_df['_mill'] != 'CRM04/06'])
                break

        # Rebuild sec_df with resolved mills
        if resolved:
            sec_df = pd.concat(resolved).drop_duplicates(
                subset=['Coil Number']).reset_index(drop=True)

        # Now group by mill and produce ONE section per mill
        mills_present = [m for m in MILL_ORDER if m in sec_df['_mill'].unique()]
        mills_present += [m for m in sec_df['_mill'].unique()
                          if m not in MILL_ORDER]

        for mill in mills_present:
            mill_df = sec_df[sec_df['_mill'] == mill]
            if mill_df.empty:
                continue
            sorted_df = sort_section(mill_df)
            out.append({
                'section_key': section_key,
                'mill':        mill,
                'label':       build_section_label(section_key, mill),
                'coils_df':    sorted_df,
            })

    return out


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
def _excel_serial(d):
    if isinstance(d, datetime):
        d = d.date()
    return (d - EXCEL_EPOCH).days


def _set_col_widths(ws):
    for i, (_, w) in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, plan_date):
    """Row 1 — banner.  Row 2 — column headers."""
    banner = f"PLANNING FOR MILL--------------AS ON  {plan_date.strftime('%d-%m-%Y')}"
    ws.cell(row=1, column=1, value=banner)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
    c = ws.cell(row=1, column=1)
    c.font = Font(name='Calibri', size=11, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color='D9E1F2')
    ws.row_dimensions[1].height = 22

    for i, (name, _) in enumerate(OUTPUT_COLUMNS, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.font = Font(name='Calibri', size=9, bold=True)
        c.alignment = Alignment(horizontal='center',
                                vertical='center',
                                wrap_text=True)
        c.fill = PatternFill('solid', start_color='F2F2F2')
        c.border = BORDER
    ws.row_dimensions[2].height = 28


def _format_data_cell(c, fmt=None):
    c.font = Font(name='Calibri', size=9)
    c.border = BORDER
    c.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        c.number_format = fmt


def _write_section_header(ws, row_idx, label, colour):
    ws.cell(row=row_idx, column=1, value=label)
    ws.merge_cells(start_row=row_idx, start_column=1,
                   end_row=row_idx,   end_column=N_COLS)
    c = ws.cell(row=row_idx, column=1)
    c.font = Font(name='Calibri', size=9, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color=colour)
    c.border = BORDER
    ws.row_dimensions[row_idx].height = 18


def _coil_row(row, plan_date_serial, customer_abbrev_lookup):
    """Build the 19-column output row for a single coil."""
    finish = _s(row.get('Surface Finish')) or _s(row.get('Edge'))
    plant = row.get('Production Plant')
    if pd.notna(plant):
        plant = int(plant)
    else:
        plant = ''

    return [
        plan_date_serial,                                  # Date
        _s(row.get('Coil Number')),                        # Batch = Coil Number
        _s(row.get('SO No')),                              # SO No
        _f(row.get('Actual Thick')),                       # Thick
        int(_f(row.get('Actual Width'))) if _f(row.get('Actual Width')) else '',
        _f(row.get('Input Coil Weight')),                  # Weight
        _f(row.get('Plan Rolling Thick 1')),               # RT
        abbreviate_customer(row.get('Customer Desc'),
                            customer_abbrev_lookup),       # Customer
        _s(row.get('Product Code')),                       # ProdCode
        _s(row.get('Actual Quality')),                     # Quality
        _s(row.get('Cust TDC')),                           # TDC
        plant,                                              # Plant
        _s(row.get('Storage Location')),                   # Storage
        _s(row.get('Planning Remark')),                    # Remark
        _s(row.get('Current Stage')),                      # Current WC
        _s(row.get('Next Stage')),                         # Next WC
        _s(row.get('Process Route')),                      # Route
        finish,                                             # Finish
        int(_f(row.get('Coil Age(# Days)'))),              # Age
    ]


_NUMBER_FORMATS = {
    3: '0.00',   # Thick
    4: '0',      # Width
    5: '0.000',  # Weight
    6: '0.00',   # RT
    11: '0',     # Plant
    18: '0',     # Age
}


def _write_coil_row(ws, row_idx, values):
    for ci, val in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        fmt = _NUMBER_FORMATS.get(ci - 1)
        _format_data_cell(c, fmt=fmt)


def _write_subtotal(ws, row_idx, start, end):
    """Subtotal row — weight column gets SUM formula."""
    for ci in range(1, N_COLS + 1):
        c = ws.cell(row=row_idx, column=ci)
        c.font = Font(name='Calibri', size=9, bold=True)
        c.fill = PatternFill('solid', start_color='F2F2F2')
        c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    f_col = get_column_letter(6)
    ws.cell(row=row_idx, column=6,
            value=f"=SUM({f_col}{start}:{f_col}{end})").number_format = '0.000'


def _write_priority_block(ws, start_row, sections):
    """
    Write the PLANNING priority block after the grand total.
    """
    crm04_sections = sorted(
        {s['section_key'] for s in sections if s['mill'] == 'CRM04'},
        key=lambda k: CRM04_PRIORITY.index(k) if k in CRM04_PRIORITY else 99,
    )
    crm06_sections = sorted(
        {s['section_key'] for s in sections if s['mill'] == 'CRM06'},
        key=lambda k: CRM06_PRIORITY.index(k) if k in CRM06_PRIORITY else 99,
    )

    r = start_row + 1

    c = ws.cell(row=r, column=1, value='PLANNING')
    c.font      = Font(name='Calibri', size=10, bold=True)
    c.alignment = Alignment(horizontal='left')
    r += 1

    for idx, sec in enumerate(crm04_sections, start=1):
        name = SECTION_SHORT_NAME.get(sec, sec)
        ws.cell(row=r, column=1, value=f"{idx}-- {name}").font = \
            Font(name='Calibri', size=9)
        if idx == 1:
            tag = ws.cell(row=r, column=7, value='CRM-04')
            tag.font = Font(name='Calibri', size=10, bold=True)
        r += 1

    r += 1
    for idx, sec in enumerate(crm06_sections, start=1):
        name = SECTION_SHORT_NAME.get(sec, sec)
        ws.cell(row=r, column=1, value=f"{idx}-- {name}").font = \
            Font(name='Calibri', size=9)
        if idx == 1:
            tag = ws.cell(row=r, column=7, value='CRM-06')
            tag.font = Font(name='Calibri', size=10, bold=True)
        r += 1


def write_sheet(wb, plan_date, sections, learning_db=None):
    """Write a single planning sheet for one date."""
    sheet_name = plan_date.strftime('%d-%m-%Y')
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    _set_col_widths(ws)
    _write_header_row(ws, plan_date)

    serial = _excel_serial(plan_date)
    cust_lookup = learning_db if learning_db else None

    row_idx = 3
    section_ranges = []  # for grand total formula
    total_coils = 0
    total_weight_sections = []

    for sec in sections:
        # Section header
        _write_section_header(ws, row_idx,
                              sec['label'],
                              SECTION_COLOURS.get(sec['section_key'], 'FFFFFF'))
        data_start = row_idx + 1

        # Data rows
        for _, coil in sec['coils_df'].iterrows():
            row_idx += 1
            vals = _coil_row(coil, serial, cust_lookup)
            _write_coil_row(ws, row_idx, vals)
            total_coils += 1

        data_end = row_idx
        # Subtotal row
        row_idx += 1
        _write_subtotal(ws, row_idx, data_start, data_end)
        section_ranges.append((data_start, data_end, row_idx))

        row_idx += 1   # ensure each section header advances correctly
        row_idx -= 0
        # leave a blank spacer row
        row_idx += 0

    # Grand total
    row_idx += 1
    for ci in range(1, N_COLS + 1):
        c = ws.cell(row=row_idx, column=ci)
        c.font = Font(name='Calibri', size=10, bold=True)
        c.fill = PatternFill('solid', start_color='FFE699')
        c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row_idx, column=1, value='GRAND TOTAL').font = \
        Font(name='Calibri', size=10, bold=True)
    if section_ranges:
        sum_parts = [f"F{s}:F{e}" for s, e, _ in section_ranges]
        ws.cell(row=row_idx, column=6,
                value=f"=SUM({','.join(sum_parts)})").number_format = '0.000'

    # Freeze panes below the header
    ws.freeze_panes = 'A3'

    # Priority block
    _write_priority_block(ws, row_idx + 2, sections)

    return total_coils, ws


# ---------------------------------------------------------------------------
# Top-level orchestrator
# ---------------------------------------------------------------------------
def generate_daily_plan(wip_file, plan_date, output_file,
                       days=1, learning_db=None, verbose=True):
    """Produce a mill plan workbook covering N consecutive days."""
    if isinstance(plan_date, str):
        plan_date = datetime.strptime(plan_date, '%Y-%m-%d').date()

    raw = load_wip(wip_file)
    original_count = len(raw)

    eligible = filter_rolling_coils(raw, learning_db)
    excluded_count = original_count - len(eligible)

    assigned = assign_all(eligible, learning_db)
    sections = build_sections(assigned, learning_db)

    wb = Workbook()
    # Remove default empty sheet
    default = wb.active
    wb.remove(default)

    summaries = []
    for d_offset in range(days):
        d = plan_date + timedelta(days=d_offset)

        # For day 2+, bump age by d_offset
        if d_offset > 0:
            day_sections = []
            for s in sections:
                df = s['coils_df'].copy()
                if 'Coil Age(# Days)' in df.columns:
                    df['Coil Age(# Days)'] = (
                        df['Coil Age(# Days)'].fillna(0).astype(float) + d_offset
                    )
                day_sections.append({**s, 'coils_df': df})
        else:
            day_sections = sections

        total_coils, _ = write_sheet(wb, d, day_sections, learning_db)

        weight_total = sum(s['coils_df']['Input Coil Weight'].sum()
                           for s in day_sections)
        summaries.append({
            'date':    d,
            'coils':   total_coils,
            'weight':  round(weight_total, 3),
            'sections': len(day_sections),
        })

    wb.save(output_file)

    if verbose:
        _print_summary(plan_date, sections, summaries[0],
                       excluded_count, raw, eligible, assigned)

    return {
        'output_file':   output_file,
        'sections':      sections,
        'summaries':     summaries,
        'eligible_count': len(eligible),
        'excluded_count': excluded_count,
    }


def _print_summary(plan_date, sections, day0_summary,
                   excluded_count, raw_df, eligible_df, assigned_df):
    print(f"\n=== MILL PLAN: {plan_date.strftime('%d-%m-%Y')} ===")
    print("Sections generated:")
    for sec in sections:
        n = len(sec['coils_df'])
        w = sec['coils_df']['Input Coil Weight'].sum()
        label = f"{sec['section_key']:25s} @ {sec['mill']:9s}"
        print(f"  {label:40s} {n:3d} coils, {w:7.1f} MT")
    print('─' * 72)
    print(f"  {'GRAND TOTAL':40s} {day0_summary['coils']:3d} coils, "
          f"{day0_summary['weight']:7.1f} MT")
    print('─' * 72)

    # Width cascade check
    cascade_ok = True
    for sec in sections:
        widths = sec['coils_df']['Actual Width'].dropna().tolist()
        if any(widths[i] < widths[i + 1] - 1
               for i in range(len(widths) - 1)):
            cascade_ok = False
    print(f"Width cascade check: {'PASS' if cascade_ok else 'WARN — non-monotonic widths in some section'}")

    # Exclusion counts
    print(f"Excluded from {len(raw_df)} raw rows: {excluded_count} "
          f"(weight<0.5, downstream, RT=0, hold, off-plant)")
    other = (assigned_df['_section'] == 'OTHER').sum()
    if other:
        print(f"⚠  {other} eligible coils could not be classified "
              f"(section=OTHER) — likely new grade/TDC combination.")
