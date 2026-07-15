"""
crm/dayplan.py — Whole-Day Rolling Sheet
==========================================
Replaces shift-wise sheets with ONE continuous, priority-ordered rolling
list per mill that covers everything planned before the next WIP refresh.

Design decisions (confirmed with planner):
  1. Per-mill, not per-shift — one long list per mill for the whole day.
  2. Rolled/pending status is PERSISTED in Supabase (same `learning_db`
     table used everywhere else) so every shift-in-charge sees the same
     live state — not just whoever has the browser tab open.
  3. Priority order is LOCKED at generation time. Ticking a coil as
     rolled never re-ranks the remaining list.
  4. All roll campaigns for the whole day are shown upfront, with the
     45-minute changeover cost applied at every roll-type transition.
  5. When a fresh WIP arrives mid-day, newly-eligible coils are APPENDED
     to the matching roll-type campaign (or a new campaign at the end)
     — existing coils and their rolled/pending status are never touched.
  6. Multiple shift-in-charges can open the same day sheet; ticking a
     coil writes straight to Supabase so the next refresh shows it.
"""
from __future__ import annotations
from datetime import date as _date, datetime
from typing import Dict, List, Optional

import pandas as pd

from . import config as C
from .scoring import SectionScore
from .campaign import alternate_order, sequence_coils

# Original WIP column names required to reproduce the standard Mill Plan
# Excel layout exactly (see generator._coil_row). Captured once per coil
# at build/append time so the export never depends on the WIP file or an
# active session still being around later.
_RAW_COLS = [
    "Coil Number", "SO No", "Actual Thick", "Actual Width",
    "Input Coil Weight", "Plan Rolling Thick 1", "Customer Desc",
    "Product Code", "Actual Quality", "Cust TDC", "Production Plant",
    "Storage Location", "Planning Remark", "Current Stage", "Next Stage",
    "Process Route", "Surface Finish", "Edge", "Coil Age(# Days)",
]


def _sequence_coils_full(df: pd.DataFrame, section_key: str) -> List[dict]:
    """
    Same ordering as campaign.sequence_coils (width↓, thick↓, grade,
    age↓) but the returned dicts carry BOTH the lean keys used by the
    scoring/campaign/UI logic (coil, mt, width, thick, rt, age, quality,
    customer, qual_risk, section) AND every raw WIP column needed to
    reproduce the standard Mill Plan Excel layout untouched.
    """
    lean = sequence_coils(df, section_key)
    if df is None or len(df) == 0:
        return lean

    # Build a lookup of raw rows keyed by (already-normalised) coil id,
    # so we can attach the raw columns to each lean dict in the same
    # width/thick/grade/age order sequence_coils already computed.
    dcol = "coil" if "coil" in df.columns else "Coil Number"
    raw_by_coil = {}
    for _, r in df.iterrows():
        cid = str(r.get(dcol, r.get("Coil Number", "")))
        raw = {}
        for col in _RAW_COLS:
            v = r.get(col, "")
            raw[col] = "" if pd.isna(v) else v
        raw_by_coil[cid] = raw

    for c in lean:
        c["raw"] = raw_by_coil.get(c["coil"], {})
    return lean

KEY_PREFIX = "dayplan_"


# ══════════════════════════════════════════════════════════════════════════════
# SERIALISATION — dataclasses  ⇄  plain JSON (Supabase-safe)
# ══════════════════════════════════════════════════════════════════════════════
def _coil_to_json(c: dict) -> dict:
    """A campaign coil dict, with rolled-tracking fields added."""
    out = dict(c)
    out.setdefault("rolled", False)
    out.setdefault("rolled_at", None)
    out.setdefault("rolled_by", None)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# BUILD — once, at day start
# ══════════════════════════════════════════════════════════════════════════════
def _new_campaign_json(mill, roll_type, section_key, consumer, coils,
                       change_from, needs_change, mt_start, reason=None):
    life = C.roll_life(roll_type, section_key, consumer)
    run  = round(sum(c["mt"] for c in coils), 2)
    mt_end = round(mt_start + run, 1)
    camp = {
        "roll_type": roll_type, "mill": mill, "sections": [section_key],
        "coils": [_coil_to_json(c) for c in coils],
        "total_mt": run, "n_coils": len(coils),
        "change_from": change_from, "needs_change": needs_change,
        "roll_life": life, "mt_start": round(mt_start, 1), "mt_end": mt_end,
        "over_life": mt_end > life, "consumer": consumer,
        "warnings": [reason] if reason else [],
    }
    return camp


def build_day_plan(
    scored:        List[SectionScore],
    mode:          str,
    current_rolls: Dict[str, str],
    mt_on_rolls:   Dict[str, float],
    n_shifts:      int = 3,
    use_alternate: Optional[Dict[str, bool]] = None,
) -> dict:
    """
    Build the WHOLE-DAY plan for both mills, once. Priority order is
    computed here and then locked — nothing downstream re-ranks it.

    Unlike the shift-wise Plan Builder, this is NOT capacity-truncated:
    every eligible coil is included. The only thing that forces a new
    campaign of the SAME roll type is the roll physically reaching its
    life limit (a dressing/change is then required, same 45-min cost as
    a roll-TYPE change). Nothing is silently dropped into "deferred".

    n_shifts is kept only as a label on the plan (informational) — it no
    longer caps how much material is included.
    use_alternate : optional {"CRM04": True/False, ...} — per mill,
                    lock in the min-changeover order instead of strict
                    priority order. Decided once, at generation time.
    """
    use_alternate = use_alternate or {}
    today = str(_date.today())
    out = {
        "date": today, "mode": mode, "n_shifts": n_shifts,
        "generated_at": datetime.utcnow().isoformat(),
        "mills": {},
    }

    for mill in ("CRM04", "CRM06"):
        ms = [s for s in scored if s.mill == mill]
        if not ms:
            continue

        roll = current_rolls.get(mill, "Light Matt")
        used = mt_on_rolls.get(mill, 0.0)

        order = (alternate_order(scored, mill, roll) if use_alternate.get(mill)
                 else sorted(ms, key=lambda x: x.rank))
        plan_type = "alternate" if use_alternate.get(mill) else "priority"

        campaigns: List[dict] = []
        n_changes = 0
        downtime  = 0

        for s in order:
            needed = s.roll_type
            seq = _sequence_coils_full(s.coils_df, s.section_key)
            if not seq:
                continue

            type_change = (needed != roll)
            old_roll = roll                    # capture BEFORE reassigning
            if type_change:
                n_changes += 1; downtime += C.ROLL_CHANGE_MIN
                used = 0.0
                roll = needed

            life = C.roll_life(needed, s.section_key, s.consumer)
            pos, first_chunk_of_section = 0, True

            # Walk this section's coils in roll-life-bounded chunks. A
            # chunk boundary mid-section means the roll physically needs
            # dressing before the rest of the section can be rolled —
            # that always costs a change, same as a roll-type change.
            while pos < len(seq):
                room = max(life - used, 0)
                chunk, run = [], 0.0
                while pos < len(seq):
                    c = seq[pos]
                    if chunk and run + c["mt"] > room + 0.05:
                        break
                    chunk.append(c); run += c["mt"]; pos += 1
                    if room > 0 and run >= room and pos < len(seq):
                        break   # room exactly filled, more coils remain

                is_change = type_change if first_chunk_of_section else True
                # change_from: the PREVIOUS roll for a genuine type
                # change (old_roll ≠ needed); the SAME roll for a
                # mid-section dressing boundary (needed == needed).
                change_src = (old_roll if (first_chunk_of_section and type_change)
                             else needed)
                reason = (None if first_chunk_of_section else
                          "🔧 Roll life reached — dressing/change required")
                camp = _new_campaign_json(
                    mill, needed, s.section_key, s.consumer, chunk,
                    change_from=change_src if is_change else "",
                    needs_change=is_change, mt_start=used, reason=reason)

                if (campaigns and campaigns[-1]["roll_type"] == needed
                        and not is_change):
                    prev = campaigns[-1]
                    prev["coils"].extend(camp["coils"])
                    if s.section_key not in prev["sections"]:
                        prev["sections"].append(s.section_key)
                    prev["total_mt"] = round(prev["total_mt"] + camp["total_mt"], 2)
                    prev["n_coils"] += camp["n_coils"]
                    prev["mt_end"]   = camp["mt_end"]
                    prev["over_life"] = prev["mt_end"] > prev["roll_life"]
                else:
                    campaigns.append(camp)

                if pos < len(seq):
                    # section not finished but the roll hit its life —
                    # next chunk starts on a freshly dressed roll
                    n_changes += 1; downtime += C.ROLL_CHANGE_MIN
                    used = 0.0
                else:
                    # section finished; MT carries forward in case the
                    # next section also needs this same roll type
                    used = camp["mt_end"]

                type_change = False
                first_chunk_of_section = False

        planned_mt = round(sum(c["total_mt"] for c in campaigns), 1)
        n_coils    = sum(c["n_coils"] for c in campaigns)

        out["mills"][mill] = {
            "mill": mill, "plan_type": plan_type, "campaigns": campaigns,
            "deferred": [], "n_changes": n_changes, "downtime_min": downtime,
            "planned_mt": planned_mt, "capacity_mt": None, "n_coils": n_coils,
            "current_roll_at_start": current_rolls.get(mill, "Light Matt"),
            "mt_on_roll_at_start":   mt_on_rolls.get(mill, 0.0),
        }
    return out


# ══════════════════════════════════════════════════════════════════════════════
# PERSISTENCE — Supabase (shared across every shift-in-charge)
# ══════════════════════════════════════════════════════════════════════════════
def _key(day: Optional[str] = None) -> str:
    from db import is_test_mode
    prefix = "TEST_" if is_test_mode() else ""
    return f"{prefix}{KEY_PREFIX}{day or str(_date.today())}"


def save_day_plan(plan: dict) -> bool:
    try:
        from db import _get_client
        client = _get_client()
        if not client:
            return False
        client.table("learning_db").upsert({
            "key":        _key(plan.get("date")),
            "data":       plan,
            "updated_at": datetime.utcnow().isoformat(),
        }).execute()
        return True
    except Exception as e:
        print(f"[dayplan] save error: {e}")
        return False


def load_day_plan(day: Optional[str] = None) -> Optional[dict]:
    try:
        from db import _get_client
        client = _get_client()
        if not client:
            return None
        resp = (client.table("learning_db")
                      .select("data")
                      .eq("key", _key(day))
                      .execute())
        if resp.data:
            return resp.data[0]["data"]
        return None
    except Exception as e:
        print(f"[dayplan] load error: {e}")
        return None


def delete_day_plan(day: Optional[str] = None) -> bool:
    try:
        from db import _get_client
        client = _get_client()
        if not client:
            return False
        client.table("learning_db").delete().eq("key", _key(day)).execute()
        return True
    except Exception as e:
        print(f"[dayplan] delete error: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
# LIVE UPDATES — tick a coil, persisted immediately for every viewer
# ══════════════════════════════════════════════════════════════════════════════
def mark_coil(plan: dict, mill: str, coil_number: str,
             rolled: bool, rolled_by: Optional[str] = None) -> dict:
    """
    Flip one coil's rolled status inside an already-loaded plan dict,
    save it back to Supabase, and return the updated plan.
    Priority order is untouched — only the coil's own flags change.
    """
    mp = plan.get("mills", {}).get(mill)
    if not mp:
        return plan
    for camp in mp["campaigns"]:
        for c in camp["coils"]:
            if c["coil"] == coil_number:
                c["rolled"]    = rolled
                c["rolled_at"] = datetime.utcnow().isoformat() if rolled else None
                c["rolled_by"] = rolled_by if rolled else None
                save_day_plan(plan)
                return plan
    return plan


def progress(plan: dict, mill: str) -> Dict:
    """Rolled vs pending MT/coils for one mill's day sheet."""
    mp = plan.get("mills", {}).get(mill)
    if not mp:
        return {"total_coils": 0, "rolled_coils": 0, "total_mt": 0.0, "rolled_mt": 0.0}
    total_coils = rolled_coils = 0
    total_mt = rolled_mt = 0.0
    for camp in mp["campaigns"]:
        for c in camp["coils"]:
            total_coils += 1
            total_mt    += c["mt"]
            if c.get("rolled"):
                rolled_coils += 1
                rolled_mt    += c["mt"]
    return {"total_coils": total_coils, "rolled_coils": rolled_coils,
            "total_mt": round(total_mt, 1), "rolled_mt": round(rolled_mt, 1)}


# ══════════════════════════════════════════════════════════════════════════════
# APPEND — new WIP arrives mid-day, existing sheet + ticks untouched
# ══════════════════════════════════════════════════════════════════════════════
def append_new_coils(plan: dict, scored: List[SectionScore]) -> Dict:
    """
    Find coils in `scored` that are NOT already anywhere in `plan`, and
    append them — respecting roll life the same way the initial build
    does (a life-exceeding append starts a fresh dressing/change chunk).

    Existing coils, their order, and their rolled status are untouched.
    Returns {"added": n, "mills": {mill: n_added}}.
    """
    existing: set = set()
    for mp in plan.get("mills", {}).values():
        for camp in mp["campaigns"]:
            for c in camp["coils"]:
                existing.add(c["coil"])

    added_total = 0
    per_mill: Dict[str, int] = {}

    for mill in ("CRM04", "CRM06"):
        mp = plan.get("mills", {}).get(mill)
        if mp is None:
            continue
        ms = [s for s in scored if s.mill == mill]
        added_here = 0

        for s in ms:
            seq = _sequence_coils_full(s.coils_df, s.section_key)
            new_coils = [c for c in seq if c["coil"] not in existing]
            if not new_coils:
                continue

            last = mp["campaigns"][-1] if mp["campaigns"] else None
            roll_before = last["roll_type"] if last else \
                mp.get("current_roll_at_start", "Light Matt")
            same_type = last is not None and last["roll_type"] == s.roll_type
            used = last["mt_end"] if same_type else 0.0
            life = C.roll_life(s.roll_type, s.section_key, s.consumer)

            pos, first_chunk = 0, True
            while pos < len(new_coils):
                room = max(life - used, 0)
                chunk, run = [], 0.0
                while pos < len(new_coils):
                    c = new_coils[pos]
                    if chunk and run + c["mt"] > room + 0.05:
                        break
                    chunk.append(c); run += c["mt"]; pos += 1
                    if room > 0 and run >= room and pos < len(new_coils):
                        break

                merges_into_last = (first_chunk and same_type)
                needs_change = not merges_into_last

                if merges_into_last:
                    last["coils"].extend(_coil_to_json(c) for c in chunk)
                    if s.section_key not in last["sections"]:
                        last["sections"].append(s.section_key)
                    last["total_mt"] = round(last["total_mt"] + run, 2)
                    last["n_coils"] += len(chunk)
                    last["mt_end"]   = round(used + run, 1)
                    last["over_life"] = last["mt_end"] > last["roll_life"]
                    mt_end = last["mt_end"]
                else:
                    reason = ("📥 Added from a later WIP refresh"
                              if first_chunk else
                              "🔧 Roll life reached — dressing/change required")
                    new_camp = _new_campaign_json(
                        mill, s.roll_type, s.section_key, s.consumer, chunk,
                        change_from=roll_before, needs_change=True,
                        mt_start=used, reason=reason)
                    mp["campaigns"].append(new_camp)
                    mp["n_changes"]    = mp.get("n_changes", 0) + 1
                    mp["downtime_min"] = mp.get("downtime_min", 0) + C.ROLL_CHANGE_MIN
                    last = new_camp
                    roll_before = s.roll_type
                    mt_end = new_camp["mt_end"]

                if pos < len(new_coils):
                    # more coils in this section, but the roll hit its
                    # life — next chunk resumes on a freshly dressed roll
                    used = 0.0
                else:
                    used = mt_end
                same_type = True   # every chunk after the first is,
                                   # by construction, the same roll type
                first_chunk = False

            added_here += len(new_coils)
            existing.update(c["coil"] for c in new_coils)

        if added_here:
            mp["planned_mt"] = round(sum(c["total_mt"] for c in mp["campaigns"]), 1)
            mp["n_coils"]    = sum(c["n_coils"] for c in mp["campaigns"])
        per_mill[mill] = added_here
        added_total   += added_here

    if added_total:
        save_day_plan(plan)
    return {"added": added_total, "mills": per_mill}


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT — exact same layout as the standard Mill Plan
# ══════════════════════════════════════════════════════════════════════════════
def _reconstruct_sections(plan: dict) -> List[dict]:
    """
    Rebuild standard {section_key, mill, label, coils_df} blocks from the
    day plan's locked-in coil order — grouped by section (not by roll
    campaign), matching how the official Mill Plan is laid out. Because
    the day plan never interleaves a section's own coils with another
    section's, every coil belonging to one section is still contiguous
    across however many roll-life chunks it was split into.
    """
    from constants import SECTION_SHORT_NAME
    sections: List[dict] = []

    for mill in ("CRM04", "CRM06"):
        mp = plan.get("mills", {}).get(mill)
        if not mp:
            continue
        order:  List[str] = []
        rows:   Dict[str, List[dict]] = {}
        for camp in mp["campaigns"]:
            for c in camp["coils"]:
                sk = c.get("section", "UNKNOWN")
                if sk not in rows:
                    rows[sk] = []
                    order.append(sk)
                rows[sk].append(c)

        for sk in order:
            recs = []
            for c in rows[sk]:
                raw = dict(c.get("raw") or {})
                if not raw:
                    # fallback for coils captured before this fix, or
                    # anything missing raw data — still exports, just
                    # with blanks in the extra WIP-only columns
                    raw = {
                        "Coil Number": c["coil"], "Actual Thick": c.get("thick", 0),
                        "Actual Width": c.get("width", 0),
                        "Input Coil Weight": c.get("mt", 0),
                        "Plan Rolling Thick 1": c.get("rt", 0),
                        "Customer Desc": c.get("customer", ""),
                        "Actual Quality": c.get("quality", ""),
                        "Coil Age(# Days)": c.get("age", 0),
                    }
                recs.append(raw)
            sections.append({
                "section_key": sk, "mill": mill,
                "label": SECTION_SHORT_NAME.get(sk, sk),
                "coils_df": pd.DataFrame(recs),
            })
    return sections


def export_day_sheet_excel(plan: dict, learning_db: Optional[dict] = None) -> bytes:
    """
    Export the day sheet in the EXACT standard Mill Plan Excel format —
    same 19-column headers, same colored section headers, same per-
    section subtotal formula, same grand total, same CRM-04/CRM-06
    priority block at the bottom (via generator.write_sheet).

    Rolled coils are highlighted with a light-green row fill and a cell
    comment showing who rolled it and when — the only addition to the
    standard layout, so the sheet stays instantly familiar to print and
    hand over, while still showing shift progress.
    """
    import io
    from datetime import date as _date_
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.comments import Comment
    from generator import write_sheet

    sections = _reconstruct_sections(plan)

    wb = Workbook()
    wb.remove(wb.active)
    plan_date = _date_.fromisoformat(plan["date"])
    _, ws = write_sheet(wb, plan_date, sections, learning_db)

    # ── Overlay rolled status — replicate write_sheet's own row math ──
    rolled_fill = PatternFill("solid", fgColor="C8E6C9")
    row = 3
    for sec in sections:
        header_row = row
        n = len(sec["coils_df"])
        data_start = header_row + 1

        # Build coil -> rolled lookup once per section from the plan
        rolled_by_coil = {}
        for camp in plan["mills"][sec["mill"]]["campaigns"]:
            for c in camp["coils"]:
                if c.get("section") == sec["section_key"]:
                    rolled_by_coil[str(c["coil"])] = c

        for i in range(n):
            r = data_start + i
            coil_no = str(sec["coils_df"].iloc[i].get("Coil Number", ""))
            info = rolled_by_coil.get(coil_no)
            if info and info.get("rolled"):
                for cell in ws[r]:
                    cell.fill = rolled_fill
                who = info.get("rolled_by") or "—"
                when = (info.get("rolled_at") or "")[:16].replace("T", " ")
                ws.cell(row=r, column=2).comment = Comment(
                    f"Rolled by {who}\n{when}", "Day Sheet")
        row = data_start + n + 1   # subtotal row
        row += 1                   # next section header row

    # ── Info sheet — day-sheet-specific context with nowhere else to go ──
    info = wb.create_sheet("Day Sheet Info")
    info.append(["Date", plan.get("date", "")])
    info.append(["Mode", plan.get("mode", "")])
    info.append(["Generated at (UTC)", plan.get("generated_at", "")])
    info.append([])
    info.append(["Mill", "Coils", "MT", "Roll Changes", "Downtime (min)"])
    for mill, mp in plan.get("mills", {}).items():
        info.append([mill, mp["n_coils"], mp["planned_mt"],
                     mp["n_changes"], mp["downtime_min"]])
    for col, w in zip("ABCDE", [22, 12, 10, 14, 16]):
        info.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()
